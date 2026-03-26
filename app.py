from flask import (Flask, render_template, request, redirect,
                   url_for, session, jsonify, flash, Response)
from werkzeug.security import check_password_hash, generate_password_hash
import sqlite3, os, functools, smtplib, threading, io
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

app = Flask(__name__)
app.secret_key = 'dainikbhaskar_portal_2025_secure'
DB_PATH = os.path.join(os.path.dirname(__file__), 'portal.db')

@app.template_filter('enumerate')
def jinja_enumerate(iterable):
    return list(enumerate(iterable))

# ── DB ────────────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def qone(conn, sql, params=()):
    row = conn.execute(sql, list(params)).fetchone()
    return row[0] if row else 0

def rows_to_dicts(rows):
    return [dict(r) for r in rows]

def and_where(w):
    return 'AND' if w else 'WHERE'

def get_setting(db, key, default=''):
    row = db.execute('SELECT value FROM app_settings WHERE key=?', (key,)).fetchone()
    return (row[0] if row else default) or default

def save_setting(db, key, value):
    db.execute("INSERT INTO app_settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=datetime('now','localtime')", (key, value))

def get_all_settings(db):
    return {r['key']: r['value'] for r in db.execute('SELECT key,value FROM app_settings').fetchall()}

def get_state_filter():
    if session.get('role') == 'engineer' and session.get('state'):
        return 'WHERE state=?', [session['state']]
    return '', []

# ── Auth ──────────────────────────────────────────────────────────────────────
def login_required(f):
    @functools.wraps(f)
    def w(*a,**k):
        if 'user_id' not in session: return redirect(url_for('login'))
        return f(*a,**k)
    return w

def admin_required(f):
    @functools.wraps(f)
    def w(*a,**k):
        if 'user_id' not in session: return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('Admin permission required','error')
            return redirect(url_for('dashboard'))
        return f(*a,**k)
    return w

def can_edit_link(link):
    """Check if current user can edit this link."""
    if session.get('role') == 'admin':
        return True
    if session.get('role') == 'engineer' and link['state'] == session.get('state'):
        return True
    return False

# ── AD Authentication ─────────────────────────────────────────────────────────
def try_ad_auth(username, password, s):
    if s.get('ad_enabled') != '1' or not s.get('ad_server','').strip():
        return False
    try:
        from ldap3 import Server, Connection, ALL, NTLM, SIMPLE, Tls
        import ssl
        use_ssl = s.get('ad_use_ssl') == '1'
        port = int(s.get('ad_port', 389))
        srv = Server(s['ad_server'].strip(), port=port, use_ssl=use_ssl, get_info=ALL)
        domain = s.get('ad_domain','').strip()
        upn = f"{username}@{domain}" if domain else username
        conn = Connection(srv, user=upn, password=password, authentication=SIMPLE)
        if conn.bind():
            conn.unbind()
            return True
        ntlm_user = f"{domain}\\{username}" if domain else username
        conn2 = Connection(srv, user=ntlm_user, password=password, authentication=NTLM)
        if conn2.bind():
            conn2.unbind()
            return True
        return False
    except Exception as e:
        print(f'AD auth error: {e}')
        return False

# ── Email ─────────────────────────────────────────────────────────────────────
def send_email(s, to_addrs, subject, html_body):
    """Send email in background thread."""
    if s.get('smtp_enabled') != '1':
        return False, 'SMTP not enabled'
    def _send():
        try:
            msg = MIMEMultipart('alternative')
            msg['Subject'] = subject
            msg['From']    = s.get('smtp_from') or s.get('smtp_user','')
            msg['To']      = ', '.join(to_addrs) if isinstance(to_addrs, list) else to_addrs
            msg.attach(MIMEText(html_body, 'html'))
            use_ssl = s.get('smtp_ssl') == '1'
            host = s.get('smtp_host','')
            port = int(s.get('smtp_port', 587))
            if use_ssl:
                srv = smtplib.SMTP_SSL(host, port, timeout=10)
            else:
                srv = smtplib.SMTP(host, port, timeout=10)
                srv.ehlo(); srv.starttls(); srv.ehlo()
            if s.get('smtp_user') and s.get('smtp_pass'):
                srv.login(s['smtp_user'], s['smtp_pass'])
            recipients = to_addrs if isinstance(to_addrs, list) else [to_addrs]
            srv.sendmail(msg['From'], recipients, msg.as_string())
            srv.quit()
        except Exception as e:
            print(f'Email error: {e}')
    threading.Thread(target=_send, daemon=True).start()
    return True, 'Sending...'

def renewal_email_body(renewals):
    rows = ''.join(f"""
    <tr style="border-bottom:1px solid #eee">
      <td style="padding:10px 14px">{r['category']}</td>
      <td style="padding:10px 14px"><b>{r['link_location'] or r['engineer_location']}</b></td>
      <td style="padding:10px 14px">{r['isp_name']}</td>
      <td style="padding:10px 14px">{r['state']}</td>
      <td style="padding:10px 14px;color:#e67e22;font-weight:700">{r['next_renewal_date']}</td>
      <td style="padding:10px 14px;color:#c0392b;font-weight:700">{r['days_left']} days</td>
      <td style="padding:10px 14px">₹{r['yearly_cost']:,.0f}</td>
    </tr>""" for r in renewals)
    return f"""
<div style="font-family:Segoe UI,sans-serif;max-width:700px;margin:0 auto">
  <div style="background:linear-gradient(135deg,#c0392b,#1e2a35);padding:24px 30px;border-radius:10px 10px 0 0">
    <h2 style="color:#fff;margin:0">📡 DB Link Management Portal</h2>
    <p style="color:rgba(255,255,255,.7);margin:4px 0 0">Renewal Alert — Action Required</p>
  </div>
  <div style="background:#fff;padding:24px 30px;border:1px solid #e2e8f0;border-top:none">
    <p style="color:#2d3748;font-size:15px">The following <b>{len(renewals)} link(s)</b> are due for renewal soon. Please process payments to avoid service interruption.</p>
    <table style="width:100%;border-collapse:collapse;font-size:13px;margin-top:16px">
      <thead>
        <tr style="background:#f7fafc">
          <th style="padding:10px 14px;text-align:left;border-bottom:2px solid #e2e8f0">Type</th>
          <th style="padding:10px 14px;text-align:left;border-bottom:2px solid #e2e8f0">Location</th>
          <th style="padding:10px 14px;text-align:left;border-bottom:2px solid #e2e8f0">ISP</th>
          <th style="padding:10px 14px;text-align:left;border-bottom:2px solid #e2e8f0">State</th>
          <th style="padding:10px 14px;text-align:left;border-bottom:2px solid #e2e8f0">Renewal Date</th>
          <th style="padding:10px 14px;text-align:left;border-bottom:2px solid #e2e8f0">Days Left</th>
          <th style="padding:10px 14px;text-align:left;border-bottom:2px solid #e2e8f0">Annual Cost</th>
        </tr>
      </thead>
      <tbody>{rows}</tbody>
    </table>
    <p style="margin-top:20px;color:#718096;font-size:12px">This is an automated alert from DB Link Management Portal. Please do not reply to this email.</p>
  </div>
  <div style="background:#f7fafc;padding:14px 30px;border-radius:0 0 10px 10px;border:1px solid #e2e8f0;border-top:none;text-align:center;font-size:11px;color:#a0aec0">
    Dainik Bhaskar IT Department — Pan India Network Management
  </div>
</div>"""

# ── ROUTES ────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return redirect(url_for('dashboard') if 'user_id' in session else url_for('login'))

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        password = request.form.get('password','')
        db = get_db()
        settings = get_all_settings(db)
        # Try AD auth first if enabled
        ad_ok = try_ad_auth(username, password, settings)
        user = db.execute('SELECT * FROM users WHERE username=?', (username,)).fetchone()
        # AD success: auto-provision user if not in DB
        if ad_ok and not user:
            db.execute("INSERT OR IGNORE INTO users (username,password_hash,role,full_name) VALUES (?,?,?,?)",
                       (username, generate_password_hash(password), 'engineer', username))
            db.commit()
            user = db.execute('SELECT * FROM users WHERE username=?', (username,)).fetchone()
        if user and (ad_ok or check_password_hash(user['password_hash'], password)):
            session.clear()
            session['user_id']   = user['id']
            session['username']  = user['username']
            session['role']      = user['role']
            session['state']     = user['state'] or ''
            session['full_name'] = user['full_name'] or user['username']
            db.close()
            return redirect(url_for('dashboard'))
        db.close()
        flash('Invalid username or password', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ── DASHBOARD ─────────────────────────────────────────────────────────────────
@app.route('/dashboard')
@login_required
def dashboard():
    db = get_db()
    where, params = get_state_filter()
    aw = and_where(where)
    stats = {
        'total':  qone(db, f'SELECT COUNT(*) FROM links {where}', params),
        'active': qone(db, f'SELECT COUNT(*) FROM links {where} {aw} link_status=?', params+['Active']),
        'cost':   qone(db, f'SELECT COALESCE(SUM(yearly_cost),0) FROM links {where}', params),
        'ill':    qone(db, f'SELECT COUNT(*) FROM links {where} {aw} category=?', params+['ILL/BB']),
        'mpls':   qone(db, f'SELECT COUNT(*) FROM links {where} {aw} category=?', params+['MPLS/PRI']),
        'p2p':    qone(db, f'SELECT COUNT(*) FROM links {where} {aw} category=?', params+['P2P']),
        'sim':    qone(db, 'SELECT COUNT(*) FROM sim_cards'),
    }
    perf_data = {r['performance']: r['cnt'] for r in
                 db.execute(f'SELECT performance, COUNT(*) as cnt FROM links {where} GROUP BY performance', params).fetchall()}
    isp_rows   = rows_to_dicts(db.execute(f'SELECT isp_name, COUNT(*) as cnt FROM links {where} GROUP BY isp_name ORDER BY cnt DESC LIMIT 10', params).fetchall())
    state_rows = rows_to_dicts(db.execute('SELECT state, COUNT(*) as cnt, COALESCE(SUM(yearly_cost),0) as cost FROM links GROUP BY state ORDER BY cnt DESC').fetchall())
    nw, np = 'WHERE is_read=0', []
    if session.get('role')=='engineer' and session.get('state'):
        nw += ' AND state=?'; np.append(session['state'])
    notifs = db.execute(f'SELECT * FROM notifications {nw} ORDER BY created_at DESC LIMIT 8', np).fetchall()
    db.close()
    return render_template('dashboard.html', stats=stats, perf_data=perf_data,
                           isp_rows=isp_rows, state_rows=state_rows, notifs=notifs)

# ── LINKS LIST ────────────────────────────────────────────────────────────────
@app.route('/links')
@login_required
def links():
    db       = get_db()
    category = request.args.get('category','')
    search   = request.args.get('search','').strip()
    state_f  = request.args.get('state','') if session['role']=='admin' else session.get('state','')
    isp_f    = request.args.get('isp','').strip()
    status_f = request.args.get('status','').strip()
    perf_f   = request.args.get('perf','').strip()
    page     = max(1, int(request.args.get('page',1) or 1))
    per_page = 30

    conds, params = [], []
    if state_f:  conds.append('state=?');        params.append(state_f)
    if category == 'ILL':
        conds.append("(category='ILL/BB' AND link_type IN ('ILL','Leased Line'))")
    elif category == 'BB':
        conds.append("(category='ILL/BB' AND link_type IN ('BB','Broadband','Airtel'))")
    elif category:
        conds.append('category=?'); params.append(category)
    if isp_f:    conds.append('isp_name=?');     params.append(isp_f)
    if status_f: conds.append('link_status=?');  params.append(status_f)
    if perf_f:   conds.append('performance=?');  params.append(perf_f)
    if search:
        conds.append('(isp_name LIKE ? OR link_location LIKE ? OR circuit_id LIKE ? OR link_ip LIKE ? OR engineer_location LIKE ? OR office_type LIKE ?)')
        params += [f'%{search}%']*6

    where  = ('WHERE '+' AND '.join(conds)) if conds else ''
    total  = qone(db, f'SELECT COUNT(*) FROM links {where}', params)
    offset = (page-1)*per_page
    link_rows = db.execute(f'SELECT * FROM links {where} ORDER BY state, link_location LIMIT ? OFFSET ?', params+[per_page,offset]).fetchall()

    all_states = db.execute('SELECT DISTINCT state FROM links WHERE state!="" ORDER BY state').fetchall()
    all_isps   = db.execute('SELECT DISTINCT isp_name FROM links WHERE isp_name!="" ORDER BY isp_name').fetchall()
    db.close()
    total_pages = max(1, (total+per_page-1)//per_page)
    return render_template('links.html', links=link_rows, total=total, page=page,
                           per_page=per_page, total_pages=total_pages, category=category,
                           search=search, state_f=state_f, isp_f=isp_f, status_f=status_f,
                           perf_f=perf_f, all_states=all_states, all_isps=all_isps)

# ── LINK DETAIL + EDIT ────────────────────────────────────────────────────────
@app.route('/links/add', methods=['GET','POST'])
@login_required
def add_link():
    if request.method == 'POST':
        db = get_db()
        try:
            cost_s = (request.form.get('yearly_cost','0') or '0').replace(',','')
            cost = float(cost_s) if cost_s.replace('.','').isdigit() else 0.0
            db.execute("""INSERT INTO links
                (category,state,engineer_location,link_location,office_type,isp_name,link_type,
                 link_status,circuit_id,media_type,speed_mbps,yearly_cost,link_ip,postal_address,
                 po_number,po_date,billing_cycle,next_renewal_date,payment_location,payment_status,remark)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
                request.form.get('category','ILL/BB'), request.form.get('state',''),
                request.form.get('engineer_location',''), request.form.get('link_location',''),
                request.form.get('office_type',''), request.form.get('isp_name',''),
                request.form.get('link_type',''), request.form.get('link_status','Active'),
                request.form.get('circuit_id',''), request.form.get('media_type',''),
                request.form.get('speed_mbps',''), cost,
                request.form.get('link_ip',''), request.form.get('postal_address',''),
                request.form.get('po_number',''), request.form.get('po_date',''),
                request.form.get('billing_cycle',''), request.form.get('next_renewal_date',''),
                request.form.get('payment_location',''), request.form.get('payment_status',''),
                request.form.get('remark','')))
            db.commit(); db.close()
            flash('New link added successfully!','success')
            return redirect(url_for('links'))
        except Exception as e:
            db.close(); flash(f'Error: {e}','error')
    db = get_db()
    all_states = db.execute('SELECT DISTINCT state FROM links WHERE state!="" ORDER BY state').fetchall()
    db.close()
    return render_template('add_link.html', all_states=all_states)

@app.route('/links/<int:lid>', methods=['GET','POST'])
@login_required
def link_detail(lid):
    db   = get_db()
    link = db.execute('SELECT * FROM links WHERE id=?', (lid,)).fetchone()
    if not link:
        db.close(); flash('Link not found','error'); return redirect(url_for('links'))
    if session['role']=='engineer' and link['state'] != session.get('state'):
        db.close(); flash('Access denied','error'); return redirect(url_for('links'))

    if request.method == 'POST':
        action = request.form.get('action','')
        if action == 'set_performance':
            perf = request.form.get('performance','Good')
            note = request.form.get('note','')
            db.execute("UPDATE links SET performance=?, updated_at=datetime('now','localtime') WHERE id=?",(perf,lid))
            db.execute("INSERT INTO link_performance (link_id,performance,note,recorded_by) VALUES (?,?,?,?)",(lid,perf,note,session['username']))
            db.commit(); flash('Performance updated!','success')
        elif action == 'edit_link' and can_edit_link(link):
            try:
                cost = float((request.form.get('yearly_cost','0') or '0').replace(',',''))
            except:
                cost = 0.0
            db.execute("""UPDATE links SET link_status=?,isp_name=?,link_type=?,speed_mbps=?,
                yearly_cost=?,next_renewal_date=?,billing_cycle=?,payment_status=?,remark=?,
                link_ip=?,circuit_id=?,media_type=?,engineer_location=?,link_location=?,
                office_type=?,postal_address=?,po_number=?,po_date=?,payment_location=?,
                updated_at=datetime('now','localtime') WHERE id=?""", (
                request.form.get('link_status',''), request.form.get('isp_name',''),
                request.form.get('link_type',''), request.form.get('speed_mbps',''),
                cost, request.form.get('next_renewal_date',''), request.form.get('billing_cycle',''),
                request.form.get('payment_status',''), request.form.get('remark',''),
                request.form.get('link_ip',''), request.form.get('circuit_id',''),
                request.form.get('media_type',''), request.form.get('engineer_location',''),
                request.form.get('link_location',''), request.form.get('office_type',''),
                request.form.get('postal_address',''), request.form.get('po_number',''),
                request.form.get('po_date',''), request.form.get('payment_location',''), lid))
            db.commit(); flash('Link updated successfully!','success')
        db.close()
        return redirect(url_for('link_detail', lid=lid))

    history = db.execute('SELECT * FROM link_performance WHERE link_id=? ORDER BY recorded_at DESC LIMIT 15',(lid,)).fetchall()
    db.close()
    return render_template('link_detail.html', link=link, history=history, can_edit=can_edit_link(link))

# ── SIM CARDS ─────────────────────────────────────────────────────────────────
@app.route('/sim-cards')
@login_required
def sim_cards():
    db       = get_db()
    search   = request.args.get('search','').strip()
    provider = request.args.get('provider','').strip()
    status_f = request.args.get('status','').strip()
    page     = max(1, int(request.args.get('page',1) or 1))
    per_page = 30
    conds, params = [], []
    if search:
        conds.append('(employee_name LIKE ? OR sim_number LIKE ? OR center LIKE ? OR service_provider LIKE ? OR wan_id LIKE ? OR employee_id LIKE ?)')
        params += [f'%{search}%']*6
    if provider: conds.append('service_provider=?'); params.append(provider)
    if status_f: conds.append('emp_status=?'); params.append(status_f)
    where  = ('WHERE '+' AND '.join(conds)) if conds else ''
    total  = qone(db, f'SELECT COUNT(*) FROM sim_cards {where}', params)
    sims   = db.execute(f'SELECT * FROM sim_cards {where} ORDER BY center LIMIT ? OFFSET ?', params+[per_page,(page-1)*per_page]).fetchall()
    providers = db.execute('SELECT DISTINCT service_provider FROM sim_cards WHERE service_provider!="" ORDER BY service_provider').fetchall()
    db.close()
    total_pages = max(1,(total+per_page-1)//per_page)
    return render_template('sim_cards.html', sims=sims, total=total, page=page,
                           per_page=per_page, total_pages=total_pages, search=search,
                           provider=provider, status_f=status_f, providers=providers)

@app.route('/sim-cards/<int:sid>', methods=['GET','POST'])
@login_required
def sim_detail(sid):
    db  = get_db()
    sim = db.execute('SELECT * FROM sim_cards WHERE id=?', (sid,)).fetchone()
    if not sim:
        db.close(); flash('SIM card not found','error'); return redirect(url_for('sim_cards'))
    if request.method == 'POST' and session['role']=='admin':
        try:
            arc = float(request.form.get('arc','0') or '0')
        except:
            arc = 0.0
        db.execute("""UPDATE sim_cards SET wan_id=?,center=?,location=?,division=?,office_type=?,
            employee_id=?,employee_name=?,sim_number=?,service_provider=?,card_type=?,
            emp_status=?,designation=?,department=?,arc=?,
            updated_at=datetime('now','localtime') WHERE id=?""", (
            request.form.get('wan_id',''), request.form.get('center',''),
            request.form.get('location',''), request.form.get('division',''),
            request.form.get('office_type',''), request.form.get('employee_id',''),
            request.form.get('employee_name',''), request.form.get('sim_number',''),
            request.form.get('service_provider',''), request.form.get('card_type',''),
            request.form.get('emp_status','Active'), request.form.get('designation',''),
            request.form.get('department',''), arc, sid))
        db.commit(); db.close(); flash('SIM card updated!','success')
        return redirect(url_for('sim_detail', sid=sid))
    db.close()
    return render_template('sim_detail.html', sim=sim)

# ── NOTIFICATIONS ─────────────────────────────────────────────────────────────
@app.route('/notifications')
@login_required
def notifications():
    db = get_db()
    conds, params = ['1=1'], []
    if session['role']=='engineer' and session.get('state'):
        conds.append('state=?'); params.append(session['state'])
    where = 'WHERE '+' AND '.join(conds)
    notifs = db.execute(f'SELECT * FROM notifications {where} ORDER BY created_at DESC', params).fetchall()
    unread = qone(db, f'SELECT COUNT(*) FROM notifications {where} AND is_read=0', params)
    db.close()
    return render_template('notifications.html', notifs=notifs, unread=unread)

@app.route('/notifications/read/<int:nid>', methods=['POST'])
@login_required
def mark_read(nid):
    db = get_db(); db.execute('UPDATE notifications SET is_read=1 WHERE id=?',(nid,)); db.commit(); db.close()
    return jsonify({'ok':True})

@app.route('/notifications/read-all', methods=['POST'])
@login_required
def mark_all_read():
    db = get_db()
    if session['role']=='engineer' and session.get('state'):
        db.execute('UPDATE notifications SET is_read=1 WHERE state=?',(session['state'],))
    else:
        db.execute('UPDATE notifications SET is_read=1')
    db.commit(); db.close()
    return redirect(url_for('notifications'))

# ── CHECK RENEWALS + EMAIL ────────────────────────────────────────────────────
@app.route('/api/check-renewals', methods=['POST'])
@login_required
def check_renewals():
    db       = get_db()
    settings = get_all_settings(db)
    days_thr = int(settings.get('notify_days','30') or 30)
    lnks     = db.execute("SELECT * FROM links WHERE next_renewal_date!=''").fetchall()
    notif_count = 0
    email_queue = []  # {state, renewals[]}
    state_renewals = {}
    today = datetime.now()

    for lnk in lnks:
        rd = (lnk['next_renewal_date'] or '').strip()
        if not rd or rd in ('nan','None'): continue
        for fmt in ('%d/%m/%Y','%Y-%m-%d','%d-%m-%Y','%m/%d/%Y'):
            try:
                renewal = datetime.strptime(rd.split(' ')[0], fmt)
                days = (renewal - today).days
                if 0 < days <= days_thr:
                    # Create portal notification
                    exists = db.execute('SELECT id FROM notifications WHERE link_id=? AND type="renewal" AND is_read=0',(lnk['id'],)).fetchone()
                    if not exists:
                        loc = lnk['link_location'] or lnk['engineer_location'] or '?'
                        db.execute('INSERT INTO notifications (link_id,message,type,state) VALUES (?,?,?,?)',(
                            lnk['id'],
                            f'Renewal in {days} days — {lnk["category"]} {loc} ({lnk["isp_name"]})',
                            'renewal', lnk['state']))
                        notif_count += 1
                    # Queue for email
                    st = lnk['state'] or 'Unknown'
                    if st not in state_renewals:
                        state_renewals[st] = []
                    state_renewals[st].append({
                        'category': lnk['category'], 'link_location': lnk['link_location'],
                        'engineer_location': lnk['engineer_location'], 'isp_name': lnk['isp_name'],
                        'state': st, 'next_renewal_date': rd, 'days_left': days,
                        'yearly_cost': lnk['yearly_cost'] or 0
                    })
                break
            except: continue

    db.commit()
    email_sent = 0
    if settings.get('smtp_enabled') == '1' and state_renewals:
        admin_email = settings.get('admin_notify_email','').strip()
        # Send per-state emails to engineers
        engineers = db.execute("SELECT * FROM users WHERE role='engineer' AND state!='' AND email!=''").fetchall()
        for eng in engineers:
            if eng['state'] in state_renewals:
                to = [eng['email']]
                if admin_email and admin_email not in to: to.append(admin_email)
                body = renewal_email_body(state_renewals[eng['state']])
                send_email(settings, to, f'[ALERT] {len(state_renewals[eng["state"]])} Link Renewals Due — {eng["state"]}', body)
                email_sent += 1
        # Send admin summary for all states
        if admin_email and state_renewals:
            all_ren = [r for rl in state_renewals.values() for r in rl]
            body = renewal_email_body(all_ren)
            send_email(settings, [admin_email], f'[SUMMARY] {len(all_ren)} Link Renewals Due — All States', body)
            email_sent += 1

    db.close()
    return jsonify({'ok':True, 'notifications': notif_count,
                    'emails_queued': email_sent,
                    'message': f'{notif_count} notifications created, {email_sent} emails queued'})

# ── REPORTS ───────────────────────────────────────────────────────────────────
@app.route('/reports')
@login_required
def reports():
    db = get_db()
    where, params = get_state_filter()
    aw = and_where(where)
    by_cat    = rows_to_dicts(db.execute(f'SELECT category, COUNT(*) as cnt, COALESCE(SUM(yearly_cost),0) as cost FROM links {where} GROUP BY category', params).fetchall())
    by_isp    = rows_to_dicts(db.execute(f'SELECT isp_name, COUNT(*) as cnt, COALESCE(SUM(yearly_cost),0) as cost FROM links {where} GROUP BY isp_name ORDER BY cnt DESC LIMIT 15', params).fetchall())
    by_perf   = rows_to_dicts(db.execute(f'SELECT performance, COUNT(*) as cnt FROM links {where} GROUP BY performance', params).fetchall())
    by_status = rows_to_dicts(db.execute(f'SELECT link_status, COUNT(*) as cnt FROM links {where} GROUP BY link_status', params).fetchall())
    by_state  = rows_to_dicts(db.execute('SELECT state, COUNT(*) as cnt, COALESCE(SUM(yearly_cost),0) as cost FROM links GROUP BY state ORDER BY cost DESC').fetchall())
    by_media  = rows_to_dicts(db.execute(f'SELECT media_type, COUNT(*) as cnt FROM links {where} {aw} media_type!="" GROUP BY media_type ORDER BY cnt DESC', params).fetchall())
    db.close()
    return render_template('reports.html', by_cat=by_cat, by_isp=by_isp, by_perf=by_perf,
                           by_status=by_status, by_state=by_state, by_media=by_media)

# ── EXPORT EXCEL ──────────────────────────────────────────────────────────────
@app.route('/export/links')
@login_required
def export_links():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    db = get_db()
    where, params = get_state_filter()
    cat = request.args.get('category','')
    aw = and_where(where)
    if cat:
        extra = f'category=?' if cat not in ('ILL','BB') else \
                f"(category='ILL/BB' AND link_type IN ('ILL','Leased Line'))" if cat=='ILL' else \
                f"(category='ILL/BB' AND link_type IN ('BB','Broadband','Airtel'))"
        where += f' {aw} {extra}'
        if cat not in ('ILL','BB'): params.append(cat)
    rows = db.execute(f'SELECT * FROM links {where} ORDER BY state, category, link_location', params).fetchall()
    db.close()

    wb = openpyxl.Workbook()
    # Sheet 1: All Links
    ws = wb.active; ws.title = 'Links'
    hdr_fill = PatternFill('solid', fgColor='C0392B')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    cols = ['ID','Category','State','Engineer Location','Link Location','Office Type','ISP','Link Type',
            'Status','Circuit ID','Media','Speed (Mbps)','Annual Cost (₹)','IP Address',
            'Billing Cycle','Renewal Date','Payment Status','Performance','Remark','Updated']
    ws.append(cols)
    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for r in rows:
        ws.append([r['id'], r['category'], r['state'], r['engineer_location'], r['link_location'],
                   r['office_type'], r['isp_name'], r['link_type'], r['link_status'], r['circuit_id'],
                   r['media_type'], r['speed_mbps'], r['yearly_cost'] or 0, r['link_ip'],
                   r['billing_cycle'], r['next_renewal_date'], r['payment_status'],
                   r['performance'], r['remark'], r['updated_at']])
    col_widths = [5,12,8,18,22,14,12,10,10,18,8,12,16,18,14,16,14,12,20,18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    # Sheet 2: Summary by State
    ws2 = wb.create_sheet('State Summary')
    ws2.append(['State','Total Links','Active Links','Annual Cost (₹)'])
    for cell in ws2[1]: cell.fill = hdr_fill; cell.font = hdr_font
    db2 = get_db()
    for row in db2.execute("SELECT state, COUNT(*) as cnt, SUM(CASE WHEN link_status='Active' THEN 1 ELSE 0 END) as act, COALESCE(SUM(yearly_cost),0) as cost FROM links GROUP BY state ORDER BY state").fetchall():
        ws2.append([row['state'], row['cnt'], row['act'], row['cost']])
    # Sheet 3: ISP Summary
    ws3 = wb.create_sheet('ISP Summary')
    ws3.append(['ISP Name','Total Links','Annual Cost (₹)','Avg Cost/Link (₹)'])
    for cell in ws3[1]: cell.fill = hdr_fill; cell.font = hdr_font
    for row in db2.execute("SELECT isp_name, COUNT(*) as cnt, COALESCE(SUM(yearly_cost),0) as cost FROM links GROUP BY isp_name ORDER BY cnt DESC").fetchall():
        avg = row['cost']/row['cnt'] if row['cnt'] else 0
        ws3.append([row['isp_name'], row['cnt'], row['cost'], round(avg)])
    db2.close()

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"DB_Links_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(buf.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename={fname}'})

@app.route('/export/sim-cards')
@login_required
def export_sim():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    db = get_db()
    rows = db.execute('SELECT * FROM sim_cards ORDER BY center').fetchall()
    db.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'SIM Cards'
    hfill = PatternFill('solid', fgColor='27AE60')
    hfont = Font(bold=True, color='FFFFFF', size=11)
    ws.append(['WAN ID','Center','Location','Division','Office Type','Employee ID','Employee Name','Designation','Department','SIM Number','Provider','Card Type','Status','Annual Cost'])
    for cell in ws[1]: cell.fill = hfill; cell.font = hfont; cell.alignment = Alignment(horizontal='center')
    for r in rows:
        ws.append([r['wan_id'],r['center'],r['location'],r['division'],r['office_type'],r['employee_id'],r['employee_name'],r['designation'],r['department'],r['sim_number'],r['service_provider'],r['card_type'],r['emp_status'],r['arc'] or 0])
    for i,w in enumerate([12,14,14,12,12,12,22,20,14,14,10,12,10,14],1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"DB_SIMCards_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(buf.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename={fname}'})

# ── USERS ─────────────────────────────────────────────────────────────────────
@app.route('/users')
@admin_required
def users():
    db = get_db()
    users = db.execute('SELECT * FROM users ORDER BY role DESC, username').fetchall()
    db.close()
    return render_template('users.html', users=users)

@app.route('/users/add', methods=['POST'])
@admin_required
def add_user():
    db = get_db()
    try:
        db.execute('INSERT INTO users (username,password_hash,role,state,full_name,email) VALUES (?,?,?,?,?,?)',(
            request.form['username'].strip(), generate_password_hash(request.form['password']),
            request.form['role'], request.form.get('state',''),
            request.form['full_name'].strip(), request.form.get('email','').strip()))
        db.commit(); flash('User added successfully!','success')
    except Exception as e:
        flash(f'Error: {e}','error')
    db.close()
    return redirect(url_for('users'))

@app.route('/users/edit/<int:uid>', methods=['GET','POST'])
@admin_required
def edit_user(uid):
    db = get_db()
    user = db.execute('SELECT * FROM users WHERE id=?',(uid,)).fetchone()
    if not user:
        db.close(); flash('User not found','error'); return redirect(url_for('users'))
    if request.method == 'POST':
        new_pass = request.form.get('new_password','').strip()
        ph = generate_password_hash(new_pass) if new_pass else user['password_hash']
        db.execute('UPDATE users SET full_name=?,email=?,role=?,state=?,password_hash=? WHERE id=?',(
            request.form['full_name'].strip(), request.form.get('email','').strip(),
            request.form['role'], request.form.get('state',''), ph, uid))
        db.commit(); db.close(); flash('User updated!','success')
        return redirect(url_for('users'))
    db.close()
    return render_template('edit_user.html', user=user)

@app.route('/users/delete/<int:uid>', methods=['POST'])
@admin_required
def delete_user(uid):
    if uid == session['user_id']:
        flash('Cannot delete your own account','error'); return redirect(url_for('users'))
    db = get_db(); db.execute('DELETE FROM users WHERE id=?',(uid,)); db.commit(); db.close()
    flash('User deleted','success'); return redirect(url_for('users'))

@app.route('/users/change-password', methods=['POST'])
@login_required
def change_password():
    old = request.form.get('old_password','')
    new = request.form.get('new_password','')
    db  = get_db()
    user = db.execute('SELECT * FROM users WHERE id=?',(session['user_id'],)).fetchone()
    if not check_password_hash(user['password_hash'], old):
        flash('Old password is incorrect','error')
    elif len(new) < 4:
        flash('Password must be at least 4 characters','error')
    else:
        db.execute('UPDATE users SET password_hash=? WHERE id=?',(generate_password_hash(new),session['user_id']))
        db.commit(); flash('Password changed successfully!','success')
    db.close()
    return redirect(url_for('users'))

# ── SETTINGS ──────────────────────────────────────────────────────────────────
@app.route('/settings', methods=['GET','POST'])
@admin_required
def settings():
    db = get_db()
    if request.method == 'POST':
        keys = ['smtp_host','smtp_port','smtp_user','smtp_pass','smtp_from','smtp_enabled','smtp_ssl',
                'ad_enabled','ad_server','ad_port','ad_base_dn','ad_domain','ad_bind_user','ad_bind_pass','ad_use_ssl',
                'notify_days','admin_notify_email']
        for k in keys:
            v = request.form.get(k,'0' if k in ('smtp_enabled','smtp_ssl','ad_enabled','ad_use_ssl') else '')
            save_setting(db, k, v)
        db.commit(); db.close(); flash('Settings saved!','success')
        return redirect(url_for('settings'))
    cfg = get_all_settings(db); db.close()
    return render_template('settings.html', cfg=cfg)

@app.route('/api/test-email', methods=['POST'])
@admin_required
def test_email():
    db = get_db(); s = get_all_settings(db); db.close()
    to = request.json.get('to','')
    if not to: return jsonify({'ok':False,'msg':'Enter a test email address'})
    ok, msg = send_email(s, [to], 'Test Email — DB Link Portal',
        '<h2>Test email from DB Link Management Portal</h2><p>SMTP configuration is working correctly!</p>')
    return jsonify({'ok':ok,'msg':msg})

@app.route('/api/test-ad', methods=['POST'])
@admin_required
def test_ad():
    db = get_db(); s = get_all_settings(db); db.close()
    uname = request.json.get('username','')
    passw = request.json.get('password','')
    if not uname: return jsonify({'ok':False,'msg':'Enter test credentials'})
    ok = try_ad_auth(uname, passw, s)
    return jsonify({'ok':ok,'msg':'AD authentication successful!' if ok else 'AD authentication failed. Check server settings.'})

if __name__ == '__main__':
    app.run(debug=True, port=5000)
